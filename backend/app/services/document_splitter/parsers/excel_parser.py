from typing import Optional

from app.services.document_splitter.models import DocumentElement


def parse_excel_elements_from_workbook(
    workbook_path: str,
    file_type: str = "xlsx",
) -> list[DocumentElement]:
    """把 Excel 工作簿解析成 table DocumentElement 列表。"""

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(
        filename=workbook_path,
        data_only=True,
        read_only=True,
    )

    elements: list[DocumentElement] = []
    source_index = 0

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            used_range = detect_used_range(worksheet)
            if used_range is None:
                continue

            min_row, max_row, min_col, max_col = used_range
            rows = extract_sheet_rows(
                worksheet,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
            table_regions = split_sheet_table_regions(rows)

            for region_index, (region_start_row, region_rows) in enumerate(table_regions):
                has_header = detect_excel_header(region_rows)
                column_count = len(region_rows[0]) if region_rows else 0
                if column_count == 0:
                    continue

                content = format_excel_region_as_markdown_table(region_rows, has_header)
                header_values = region_rows[0] if has_header else []
                data_row_start = region_start_row + (1 if has_header else 0)
                data_row_end = region_start_row + len(region_rows) - 1

                metadata = {
                    "file_type": file_type,
                    "sheet_name": worksheet.title,
                    "heading_path": [worksheet.title],
                    "block_type": "table",
                    "splitter": "excel_table_block",
                    "source_parser": "excel_parser",
                    "has_header": has_header,
                    "header_row_count": 1 if has_header else 0,
                    "header_values": header_values,
                    "column_count": column_count,
                    "col_start": column_index_to_name(min_col - 1),
                    "col_end": column_index_to_name(max_col - 1),
                    "table_region_index": region_index,
                    "sheet_index": sheet_index,
                    "sheet_used_range": build_sheet_range_label(min_row, max_row, min_col, max_col),
                    "table_format": "excel_markdown",
                }
                if has_header:
                    metadata["header_row"] = region_start_row
                if len(region_rows) > (1 if has_header else 0):
                    metadata["row_start"] = data_row_start
                    metadata["row_end"] = data_row_end
                    metadata["row_count"] = data_row_end - data_row_start + 1

                elements.append(
                    DocumentElement(
                        source_index=source_index,
                        element_type="table",
                        text=content,
                        sheet_name=worksheet.title,
                        row_start=metadata.get("row_start"),
                        row_end=metadata.get("row_end"),
                        col_start=metadata["col_start"],
                        col_end=metadata["col_end"],
                        metadata=metadata,
                    )
                )
                source_index += 1
    finally:
        workbook.close()

    return elements


def workbook_to_text(workbook_path: str) -> str:
    """把 Excel 工作簿转成可读的纯文本摘要，供 documents.extracted_text 保存。"""

    openpyxl = load_openpyxl()
    workbook = openpyxl.load_workbook(
        filename=workbook_path,
        data_only=True,
        read_only=True,
    )
    parts: list[str] = []

    try:
        for worksheet in workbook.worksheets:
            used_range = detect_used_range(worksheet)
            if used_range is None:
                continue

            min_row, max_row, min_col, max_col = used_range
            rows = extract_sheet_rows(
                worksheet,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
            if not rows:
                continue

            parts.append(f"Sheet: {worksheet.title}")
            for _, row_values in rows:
                parts.append("\t".join(row_values).rstrip())
            parts.append("")
    finally:
        workbook.close()

    return "\n".join(parts).strip()


def load_openpyxl():
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel parsing. Install it with `pip install openpyxl`."
        ) from exc
    return openpyxl


def detect_used_range(worksheet) -> Optional[tuple[int, int, int, int]]:
    """检测 sheet 的 used range，并裁掉外围空白。"""

    min_row: Optional[int] = None
    max_row: Optional[int] = None
    min_col: Optional[int] = None
    max_col: Optional[int] = None

    for row in worksheet.iter_rows():
        for cell in row:
            if is_empty_excel_value(cell.value):
                continue
            if min_row is None or cell.row < min_row:
                min_row = cell.row
            if max_row is None or cell.row > max_row:
                max_row = cell.row
            if min_col is None or cell.column < min_col:
                min_col = cell.column
            if max_col is None or cell.column > max_col:
                max_col = cell.column

    if None in (min_row, max_row, min_col, max_col):
        return None

    return (min_row, max_row, min_col, max_col)


def extract_sheet_rows(
    worksheet,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[tuple[int, list[str]]]:
    """按 used range 抽取 sheet 的行数据。"""

    rows: list[tuple[int, list[str]]] = []

    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=min_row,
    ):
        values = [normalize_excel_cell(value) for value in row]
        rows.append((row_index, values))

    return rows


def split_sheet_table_regions(
    rows: list[tuple[int, list[str]]],
) -> list[tuple[int, list[list[str]]]]:
    """按空行拆出 sheet 里的多个纵向 table region。"""

    regions: list[tuple[int, list[list[str]]]] = []
    current_region: list[list[str]] = []
    current_region_start: Optional[int] = None

    for row_index, values in rows:
        if is_blank_values_row(values):
            if current_region:
                regions.append((current_region_start or row_index, current_region))
                current_region = []
                current_region_start = None
            continue

        if current_region_start is None:
            current_region_start = row_index
        current_region.append(values)

    if current_region:
        regions.append((current_region_start or 1, current_region))

    return regions


def detect_excel_header(rows: list[list[str]]) -> bool:
    """基础 header 检测。"""

    if len(rows) < 2:
        return False

    first_row = rows[0]
    second_row = rows[1]

    score = 0
    if any(cell for cell in first_row):
        score += 1
    if not any(is_probably_numeric(cell) for cell in first_row) and any(
        is_probably_numeric(cell) for cell in second_row
    ):
        score += 1
    if first_row != second_row:
        score += 1

    return score >= 2


def format_excel_region_as_markdown_table(rows: list[list[str]], has_header: bool) -> str:
    if not rows:
        return ""

    lines: list[str] = []
    start_row_index = 1 if has_header else 0

    if has_header:
        lines.append(format_markdown_table_row(rows[0]))
        lines.append(format_markdown_table_separator(len(rows[0])))

    for row in rows[start_row_index:]:
        lines.append(format_markdown_table_row(row))

    if not lines:
        lines.append(format_markdown_table_row(rows[0]))

    return "\n".join(lines).strip()


def format_markdown_table_row(cells: list[str]) -> str:
    return f"| {' | '.join(escape_markdown_table_cell(cell) for cell in cells)} |"


def format_markdown_table_separator(column_count: int) -> str:
    return f"| {' | '.join(['---'] * column_count)} |"


def escape_markdown_table_cell(cell: str) -> str:
    return cell.replace("\n", " ").replace("|", "\\|").strip()


def normalize_excel_cell(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def is_empty_excel_value(value) -> bool:
    return value is None or not str(value).strip()


def is_blank_values_row(values: list[str]) -> bool:
    return not any(value.strip() for value in values)


def is_probably_numeric(value: str) -> bool:
    if not value:
        return False
    stripped_value = value.replace(",", "").replace("%", "")
    try:
        float(stripped_value)
        return True
    except ValueError:
        return False


def column_index_to_name(index: int) -> str:
    if index < 0:
        return "A"

    result = ""
    current = index + 1
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def build_sheet_range_label(min_row: int, max_row: int, min_col: int, max_col: int) -> str:
    return (
        f"{column_index_to_name(min_col - 1)}{min_row}:"
        f"{column_index_to_name(max_col - 1)}{max_row}"
    )
